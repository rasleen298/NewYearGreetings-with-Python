#Section-1:  __Creating a workbook__
'''
import openpyxl
filepath='/Users/rasleen/Documents/mtexcel_file.xlsx'

wb=openpyxl.Workbook()
#Getting a sheet to work with
ws = wb.active
#Assign Values to cell
ws["A1"] = "Name"
ws["B1"] = "Age"
ws["A2"] = "Paul"
ws["B2"] = 28

wb.save(filepath)
'''
#Section-2:  __Creating new sheets__

wb = openpyxl.load_workbook(filepath)
 
sheet_two = wb.create_sheet("Email details")

 ws = wb["Email details"]
 
#Creating cell values
data_frame = [
    ["Name",'Age',"Email"],
    ["Godson","30","navika7244+person4@gmail.com"],
    ["Esther","27","navika7244+person1@gmail.com"],
    ["Frank","33","rasleen0209@gmail.com"],
    ["Sandra","30","navika7244+person3@gmail.com"]
]
 
#Appending cell values to cell
for data in data_frame:
    ws.append(data)
 
#Saving workbook
wb.save(filepath)