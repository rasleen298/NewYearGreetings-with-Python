#A Python Script to send emails to addresses stored in an Excel file
import openpyxl
import smtplib
import getpass

file_path = "/Users/rasleen/Documents/mtexcel_file.xlsx"
wb = openpyxl.load_workbook(file_path)
 
#Get sheet where you have stored the Contact details

sheet = wb["Email details"]
 
names = [sheet.cell(row=i,column=1).value for i in range(2,6)]
age = [sheet.cell(row=i,column=2).value for i in range(2,6)]
email = [sheet.cell(row=i,column=3).value for i in range(2,6)]
print(email)
#message to send to clients
message = "Subject: 2019 greetings\n\nHello {0},\n Wishing you a year full of happiness,joy and good wealth\n"
 
host = "smtp.gmail.com"
port = 587
username = input("Enter your email")
password = getpass.getpass("Type your password and press enter: ",stream=None)

s = smtplib.SMTP(host,port)
s.ehlo()
s.starttls()
s.login(username,password)
i = 0
 
while i < len(email):
    for name in names:
        if name==None:
            pass
        else:
            s.sendmail(username,email[i],message.format(name))  
        i+=1
        
s.quit()
